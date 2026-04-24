import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import os
#TEST#
def create_control_panel():
    wb = openpyxl.Workbook()
    script_dir = os.path.dirname(os.path.abspath(__file__))
    full_path = os.path.join(script_dir, "simulation_inputs.xlsx")

    # --- STYLING CONSTANTS ---
    HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    WHITE_FONT = Font(bold=True, color="FFFFFF")
    BOLD_FONT = Font(bold=True)

    def add_section(ws, row, title):
        cell = ws.cell(row=row, column=1, value=title)
        cell.font = WHITE_FONT
        cell.fill = HEADER_FILL
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        return row + 1

    def add_var(ws, row, name, value, desc):
        ws.cell(row=row, column=1, value=name).font = BOLD_FONT
        ws.cell(row=row, column=2, value=value)
        ws.cell(row=row, column=3, value=desc)
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 55
        return row + 1

    # ── SHEET 1: MAIN CONTROL PANEL (Strategic Levers) ───────────────────────
    ws_front = wb.active
    ws_front.title = "Main Control Panel"
    row = 1

    row = add_section(ws_front, row, "1. Scenario Metadata")
    row = add_var(ws_front, row, "SCENARIO_NAME", "Industrial_V5_Full", "Suffix for results")
    row = add_var(ws_front, row, "NUM_JOBS", 5000, "Recommended for 3rd bucket significance")
    row = add_var(ws_front, row, "START_DATE", "2022-01-01", "Simulation start (YYYY-MM-DD)")
    row = row + 1

    row = add_section(ws_front, row, "2. Press Fleet Job Shares (Must sum to 1.0)")
    row = add_var(ws_front, row, "SHARE_2190", 0.30, "2190 KBA106 — Perfecting, fleet workhorse")
    row = add_var(ws_front, row, "SHARE_2160", 0.21, "2160 840 Komori — Sheetfed, short-run specialist")
    row = add_var(ws_front, row, "SHARE_2150", 0.19, "2150 640 Komori — Sheetfed")
    row = add_var(ws_front, row, "SHARE_2500", 0.16, "2500 640 Komori — Sheetfed")
    row = add_var(ws_front, row, "SHARE_2330", 0.08, "2330 640 Komori — Sheetfed")
    row = add_var(ws_front, row, "SHARE_2060", 0.06, "2060 KBA105 — Perfecting, most unreliable")
    row = row + 1

    row = add_section(ws_front, row, "3. Shift Mix (Must sum to 1.0)")
    row = add_var(ws_front, row, "SHIFT_MIX_RED_DAY", 0.25, "Target % for Red Day")
    row = add_var(ws_front, row, "SHIFT_MIX_RED_NIGHT", 0.25, "Target % for Red Night")
    row = add_var(ws_front, row, "SHIFT_MIX_BLACK_DAY", 0.25, "Target % for Black Day")
    row = add_var(ws_front, row, "SHIFT_MIX_BLACK_NIGHT", 0.25, "Target % for Black Night")
    row = row + 1

    row = add_section(ws_front, row, "4. Layout Frequency (100-up = Foil | 121-up = White)")
    row = add_var(ws_front, row, "MIX_121_PLAIN", 0.40, "Frequency of 121-up White stock")
    row = add_var(ws_front, row, "MIX_121_HOLO", 0.25, "Frequency of 121-up Foil stock")
    row = add_var(ws_front, row, "MIX_100_PLAIN", 0.15, "Frequency of 100-up White stock")
    row = add_var(ws_front, row, "MIX_100_HOLO", 0.20, "Frequency of 100-up Foil stock")
    row = row + 1

    row = add_section(ws_front, row, "5. Strategic Profitability Levers")
    row = add_var(ws_front, row, "CUST_A_MARKUP", 1.25, "Base markup for Major Account")
    row = add_var(ws_front, row, "CUST_B_MARKUP", 1.40, "Base markup for Secondary Account")
    row = add_var(ws_front, row, "SPOT_MARKUP_PREMIUM", 1.55, "Base markup for Spot Jobs")
    row = add_var(ws_front, row, "COMPLEXITY_PREMIUM_FOIL", 0.15, "Additional markup nudge for Foil")
    row = row + 1

    # ── SHEET 2: BACKEND ENGINE (Manufacturing Constants) ────────────────────
    ws_back = wb.create_sheet("Backend Engine")
    row = 1

    row = add_section(ws_back, row, "1. Equipment Age Penalties (1.0 = new — no press in this fleet is below 1.15)")
    row = add_var(ws_back, row, "AGE_FACTOR_2190", 1.15, "2190 KBA106 — Perfecting, best maintained")
    row = add_var(ws_back, row, "AGE_FACTOR_2500", 1.20, "2500 640 Komori — Sheetfed, solid throughput")
    row = add_var(ws_back, row, "AGE_FACTOR_2330", 1.20, "2330 640 Komori — Sheetfed, consistent")
    row = add_var(ws_back, row, "AGE_FACTOR_2150", 1.25, "2150 640 Komori — Sheetfed, middle of road")
    row = add_var(ws_back, row, "AGE_FACTOR_2160", 1.30, "2160 840 Komori — Sheetfed, heavy short-run use")
    row = add_var(ws_back, row, "AGE_FACTOR_2060", 1.50, "2060 KBA105 — Perfecting, most unreliable")
    row = row + 1

    row = add_section(ws_back, row, "2. Financials & Costs")
    row = add_var(ws_back, row, "SHEETFED_RATE_HR",      110, "Actual cost to run SF press per hour (fully loaded)")
    row = add_var(ws_back, row, "PERFECTING_RATE_HR",    125, "Actual cost to run PF press per hour (fully loaded)")
    row = add_var(ws_back, row, "SHEETFED_BILL_RATE_HR", 250, "Hourly rate billed to customer for SF press time")
    row = add_var(ws_back, row, "PERFECTING_BILL_RATE_HR",285, "Hourly rate billed to customer for PF press time")
    row = add_var(ws_back, row, "STOCK_COST_WHITE", 55,  "Cost per MSF White")
    row = add_var(ws_back, row, "STOCK_COST_FOIL",  320, "Cost per MSF Foil")
    row = add_var(ws_back, row, "INK_COST_PER_LB",  20,  "Cost of ink per lb")
    row = row + 1

    row = add_section(ws_back, row, "3. Base Speeds & Output Math")
    row = add_var(ws_back, row, "PERFECTING_MAKEREADY_BONUS", 0.8, "MR time reduction for PF")
    row = add_var(ws_back, row, "BASE_SPEED_WHITE_SHEETFED", 10500, "Baseline SPH White SF")
    row = add_var(ws_back, row, "BASE_SPEED_WHITE_PERFECTING", 9500, "Baseline SPH White PF")
    row = add_var(ws_back, row, "BASE_SPEED_FOIL_SHEETFED", 7500, "Baseline SPH Foil SF")
    row = add_var(ws_back, row, "BASE_SPEED_FOIL_PERFECTING", 6500, "Baseline SPH Foil PF")
    row = row + 1

    row = add_section(ws_back, row, "4. Hard QC Failure Thresholds (Defect Triggers)")
    row = add_var(ws_back, row, "DELTA_E_REJECT", 3.5, "Max color drift tolerance")
    row = add_var(ws_back, row, "REGISTER_REJECT", 2, "Max misreg (thousandths)")
    row = add_var(ws_back, row, "DOT_GAIN_REJECT", 30, "Max allowable dot gain %")
    row = add_var(ws_back, row, "CUT_DEVIATION_REJECT", 0.5, "Max cut deviation (mm)")
    row = add_var(ws_back, row, "FOIL_ADHESION_FAIL", 70, "Minimum required foil adhesion score (0-100)")
    row = row + 1

    row = add_section(ws_back, row, "5. Continuous Waste & Noise Physics")
    row = add_var(ws_back, row, "NIGHT_WASTE_FACTOR", 1.15, "Continuous waste multiplier for night shift")
    row = add_var(ws_back, row, "NIGHT_QUALITY_FACTOR", 1.15, "QC error multiplier for night shift")
    row = add_var(ws_back, row, "FOIL_WASTE_FACTOR", 1.25, "Extra waste for Foil stock (also drives jam rate)")
    row = add_var(ws_back, row, "SPEED_NOISE_STD", 400, "Std Dev for running speeds")
    row = add_var(ws_back, row, "INK_COVERAGE_MIN", 0.55, "Min ink coverage per sheet")
    row = add_var(ws_back, row, "INK_COVERAGE_MAX", 0.90, "Max ink coverage per sheet")
    row = add_var(ws_back, row, "DEFECT_WINDOW_SHEETS", 50, "Sheets tossed on a QC hit (readjust and keep running)")
    row = add_var(ws_back, row, "QC_READJUST_MINUTES", 15, "Press downtime to readjust after a QC hit")
    row = add_var(ws_back, row, "MAKEREADY_NOISE_STD", 12, "Std dev for setup time (minutes)")
    row = row + 1

    row = add_section(ws_back, row, "6. Jam Model (Discrete Mechanical Events)")
    row = add_var(ws_back, row, "JAM_RATE_PER_10K_SHEETS", 0.03, "Poisson rate per 10K sheets run (pre-age/foil)")
    row = add_var(ws_back, row, "JAM_MINUTES", 20, "Press downtime in minutes per jam")
    row = add_var(ws_back, row, "JAM_WASTE_SHEETS", 30, "Sheets scrapped per jam event")
    row = row + 1

    row = add_section(ws_back, row, "7. Delivery Time Physics")
    row = add_var(ws_back, row, "LEAD_TIME_MEAN_DAYS", 8,   "Avg days quoted to customer")
    row = add_var(ws_back, row, "LEAD_TIME_STD_DAYS",  2,   "Std dev of quoted lead time")
    row = add_var(ws_back, row, "DELIVERY_BASE_DAYS",  3,   "Baseline prep + ship days (24hr plant)")
    row = add_var(ws_back, row, "DELIVERY_HOURS_PER_DAY", 24.0, "Plant runs 24hrs — press hrs to day conversion")
    row = add_var(ws_back, row, "ACTUAL_TIME_STD",     0.75, "Std dev of noise on computed delivery (tight — 24hr plant)")
    row = row + 1

    row = add_section(ws_back, row, "8. General Volume & Customer Shares")
    row = add_var(ws_back, row, "CUST_A_SHARE", 0.70, "Major Account")
    row = add_var(ws_back, row, "CUST_B_SHARE", 0.20, "Secondary Account")
    row = add_var(ws_back, row, "SPOT_JOB_SHARE", 0.10, "High-margin Spot Jobs")
    row = add_var(ws_back, row, "RANDOM_SEED", 42, "Seed for reproducible scenarios")
    row = row + 1

    row = add_section(ws_back, row, "9. Complex Matrices (Hardcoded Pointers)")
    row = add_var(ws_back, row, "INK_CONFIG_MIX", "In Code: Sec 3", "Probability of configurations (4/4, UV, etc.)")
    row = add_var(ws_back, row, "WASTE_BASE", "In Code: Sec 3", "Matrix of baseline waste percentages")
    row = add_var(ws_back, row, "QTY_OPTIONS", "In Code: Sec 3", "Distribution of order quantities")
    row = add_var(ws_back, row, "PLATE_COST", "In Code: Sec 3", "Setup costs for plates per config")

    wb.save(full_path)
    return full_path

if __name__ == "__main__":
    path = create_control_panel()
    print(f"✅ V5 Dashboard Created (Frontend + Backend) at: {path}")
