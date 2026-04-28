"""
Trading Card Print Shop — Synthetic Dataset Generator
======================================================
All scenario variables are in the CONFIG section below.
Tweak anything there and re-run to generate a new dataset.

Output: trading_card_print_data.csv  (and optionally .xlsx)

Requires: pandas, numpy, openpyxl
Run with: python trading_card_generate_dataset.py
"""

import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import random

# ══════════════════════════════════════════════════════════════════════════════
# CONFIG — Edit anything in this section to change the scenario
# ══════════════════════════════════════════════════════════════════════════════

# ── General ───────────────────────────────────────────────────────────────────
RANDOM_SEED     = 42            # Change for a different random dataset
NUM_JOBS        = 1000          # Number of print jobs to generate
START_DATE      = "2022-01-01"  # First job date (YYYY-MM-DD)
DATE_RANGE_DAYS = 730           # Spread jobs over this many days (~2 years)

OUTPUT_CSV      = "trading_card_print_data.csv"
OUTPUT_XLSX     = "trading_card_print_data.xlsx"  # Set to None to skip Excel
EXPORT_EXCEL    = True

# ── Shifts ────────────────────────────────────────────────────────────────────
# Must sum to 1.0
SHIFTS = {
    "Red Day":    0.25,
    "Red Night":  0.25,
    "Black Day":  0.25,
    "Black Night": 0.25,
}

# Night shift multiplier on waste and quality errors (1.0 = no effect)
NIGHT_WASTE_FACTOR   = 1.15
NIGHT_QUALITY_FACTOR = 1.15

# ── Presses ───────────────────────────────────────────────────────────────────
# press_name: (type, job_share, age_factor)
#   type        — "Sheetfed" or "Perfecting"
#   job_share   — proportion of jobs run on this press (must sum to 1.0)
#   age_factor  — >1.0 means older/worse press (higher waste, slower speed)
PRESSES = {
    "SF-1": ("Sheetfed",   0.20, 1.0),
    "SF-2": ("Sheetfed",   0.20, 1.0),
    "SF-3": ("Sheetfed",   0.18, 1.2),   # older press
    "PF-1": ("Perfecting", 0.22, 1.0),
    "PF-2": ("Perfecting", 0.20, 1.2),   # older press
}

# Press hourly rates ($/hr)
SHEETFED_RATE_HR   = 295
PERFECTING_RATE_HR = 340  # higher — more complex machine

# ── Stock ─────────────────────────────────────────────────────────────────────
# Must sum to 1.0
STOCK_MIX = {
    "White": 0.65,
    "Foil":  0.35,
}

# Cost per thousand sheets (MSF)
STOCK_COST_PER_MSF = {
    "White": 55,
    "Foil":  320,
}

# Foil multiplier on waste and makeready (foil is harder to run)
FOIL_WASTE_FACTOR     = 1.25
FOIL_MAKEREADY_FACTOR = 1.25
FOIL_SPEED_PENALTY    = 0.70  # foil presses run at 70% of white speed

# ── Ink Configurations ────────────────────────────────────────────────────────
# Must sum to 1.0. Foil jobs have a 50% chance of getting Foil Stamp regardless.
INK_CONFIG_MIX = {
    "4/4 CMYK":        0.45,
    "4/4 + Spot Gloss": 0.25,
    "4/4 + Foil Stamp": 0.15,
    "4/4 + Spot UV":    0.15,
}

# Plate cost by ink config ($)
PLATE_COST = {
    "4/4 CMYK":        260,
    "4/4 + Spot Gloss": 380,
    "4/4 + Foil Stamp": 520,
    "4/4 + Spot UV":    380,
}

# Finishing cost per sheet ($) — 0 means no finishing cost
FINISHING_COST_PER_SHEET = {
    "4/4 CMYK":        0.000,
    "4/4 + Spot Gloss": 0.009,
    "4/4 + Foil Stamp": 0.018,
    "4/4 + Spot UV":    0.009,
}

# ── Job Specs ─────────────────────────────────────────────────────────────────
# Cards per sheet layouts and their frequency
CARDS_PER_SHEET_OPTIONS = {100: 0.30, 121: 0.7}

# Quantity ordered (sheets) and frequency
QTY_OPTIONS = {
    500:   0.05,
    1000:  0.10,
    2000:  0.20,
    3000:  0.20,
    5000:  0.25,
    7500:  0.12,
    10000: 0.08,
}

NUM_CUSTOMERS  = 3   # Unique customer pool size
NUM_CARD_SETS  = 40   # Unique card set pool size

# ── Press Performance ─────────────────────────────────────────────────────────
# Base makeready time (minutes) by ink config
MAKEREADY_BASE_MIN = {
    "4/4 CMYK":        75,
    "4/4 + Spot Gloss": 90,
    "4/4 + Foil Stamp": 120,
    "4/4 + Spot UV":    95,
}

# Perfecting presses save makeready time (no re-feed second pass)
PERFECTING_MAKEREADY_BONUS = 0.80  # 20% faster makeready on perfectors

# Base press speed (sheets/hr) — before foil and age adjustments
BASE_SPEED_WHITE_SHEETFED    = 10500
BASE_SPEED_WHITE_PERFECTING  = 9500
BASE_SPEED_FOIL_SHEETFED     = 7500
BASE_SPEED_FOIL_PERFECTING   = 6500

# Makeready time noise (std dev in minutes)
MAKEREADY_NOISE_STD = 12

# Press speed noise (std dev in sph)
SPEED_NOISE_STD = 400

# ── Waste ─────────────────────────────────────────────────────────────────────
# Base waste % by stock and ink
WASTE_BASE = {
    ("White", "4/4 CMYK"):        0.025,
    ("White", "4/4 + Spot Gloss"): 0.038,
    ("White", "4/4 + Foil Stamp"): 0.045,
    ("White", "4/4 + Spot UV"):    0.038,
    ("Foil",  "4/4 CMYK"):        0.055,
    ("Foil",  "4/4 + Spot Gloss"): 0.060,
    ("Foil",  "4/4 + Foil Stamp"): 0.065,
    ("Foil",  "4/4 + Spot UV"):    0.060,
}
WASTE_NOISE_STD = 0.008

# ── Quality Thresholds ────────────────────────────────────────────────────────
# Jobs failing ANY threshold are flagged quality_pass = 0
DELTA_E_REJECT       = 3.5   # Color accuracy (lower = better)
REGISTER_REJECT      = 2.0   # Misregister in thousandths of inch
DOT_GAIN_REJECT      = 30.0  # Dot gain % upper limit
CUT_DEVIATION_REJECT = 0.5   # Card cut deviation mm upper limit
FOIL_ADHESION_FAIL   = 70    # Foil adhesion score lower limit (foil only)

# Base delta E by stock (exponential distribution scale)
DELTA_E_BASE = {"White": 1.2, "Foil": 1.8}

# Base register error by stock
REGISTER_BASE = {"White": 0.5, "Foil": 0.8}

# Probability that a failed job requires a rerun (vs. salvaged/written off)
RERUN_PROBABILITY = 0.75

# ── Financials ────────────────────────────────────────────────────────────────
INK_COST_PER_LB    = 20     # $/lb
INK_COVERAGE_RANGE = (0.55, 0.90)  # Min/max ink coverage % (trading cards are vivid)

# Revenue markup range over total cost (uniform distribution)
MARKUP_MIN = 1.18
MARKUP_MAX = 1.65

# ── Delivery ──────────────────────────────────────────────────────────────────
LEAD_TIME_MEAN_DAYS = 8   # Average days from job date to due date
LEAD_TIME_STD_DAYS  = 2
ACTUAL_TIME_MEAN    = 8   # Average actual delivery days
ACTUAL_TIME_STD     = 2

# ══════════════════════════════════════════════════════════════════════════════
# GENERATION — No need to edit below this line
# ══════════════════════════════════════════════════════════════════════════════

def generate_dataset():
    np.random.seed(RANDOM_SEED)
    random.seed(RANDOM_SEED)

    n = NUM_JOBS
    start = datetime.strptime(START_DATE, "%Y-%m-%d")
    dates = sorted([start + timedelta(days=random.randint(0, DATE_RANGE_DAYS)) for _ in range(n)])

    # Unpack press config
    press_names  = list(PRESSES.keys())
    press_shares = [v[1] for v in PRESSES.values()]
    press_type_map  = {k: v[0] for k, v in PRESSES.items()}
    press_age_map   = {k: v[2] for k, v in PRESSES.items()}

    # Sample columns
    shifts     = np.random.choice(list(SHIFTS.keys()),    n, p=list(SHIFTS.values()))
    presses    = np.random.choice(press_names,             n, p=press_shares)
    stock_type = np.random.choice(list(STOCK_MIX.keys()), n, p=list(STOCK_MIX.values()))
    ink_config = np.random.choice(list(INK_CONFIG_MIX.keys()), n, p=list(INK_CONFIG_MIX.values()))
    cards_per_sheet = np.random.choice(
        list(CARDS_PER_SHEET_OPTIONS.keys()), n, p=list(CARDS_PER_SHEET_OPTIONS.values()))
    qty_sheets = np.random.choice(
        list(QTY_OPTIONS.keys()), n, p=list(QTY_OPTIONS.values()))

    customers  = [f"CUST-{str(random.randint(1, NUM_CUSTOMERS)).zfill(3)}" for _ in range(n)]
    card_sets  = [f"SET-{str(random.randint(1, NUM_CARD_SETS)).zfill(3)}"  for _ in range(n)]

    # Foil jobs get Foil Stamp 50% of the time
    ink_config = np.where(
        (stock_type == "Foil") & (np.random.rand(n) < 0.50),
        "4/4 + Foil Stamp", ink_config)

    # Derived arrays
    press_type_arr = np.array([press_type_map[p] for p in presses])
    press_age_arr  = np.array([press_age_map[p]  for p in presses])
    is_night       = np.isin(shifts, ["Red Night", "Black Night"])
    is_foil        = stock_type == "Foil"
    is_perfecting  = press_type_arr == "Perfecting"

    night_factor = np.where(is_night, NIGHT_WASTE_FACTOR, 1.0)
    foil_factor  = np.where(is_foil,  FOIL_WASTE_FACTOR,  1.0)

    # Makeready
    makeready_base = np.array([MAKEREADY_BASE_MIN[c] for c in ink_config], dtype=float)
    makeready_time = (
        makeready_base
        * np.where(is_foil, FOIL_MAKEREADY_FACTOR, 1.0)
        * np.where(is_night, NIGHT_QUALITY_FACTOR, 1.0)
        * press_age_arr
        * np.where(is_perfecting, PERFECTING_MAKEREADY_BONUS, 1.0)
        + np.random.normal(0, MAKEREADY_NOISE_STD, n)
    ).clip(25, 200).astype(int)

    # Press speed
    speed_base = np.where(
        is_foil & is_perfecting, BASE_SPEED_FOIL_PERFECTING,
        np.where(is_foil,        BASE_SPEED_FOIL_SHEETFED,
        np.where(is_perfecting,  BASE_SPEED_WHITE_PERFECTING,
                                 BASE_SPEED_WHITE_SHEETFED)))
    press_speed = (speed_base / press_age_arr
                   + np.random.normal(0, SPEED_NOISE_STD, n)).clip(4000, 13000).astype(int)

    # Waste
    waste_base_arr = np.array([
        WASTE_BASE.get((s, c), 0.04) for s, c in zip(stock_type, ink_config)], dtype=float)
    waste_pct = (waste_base_arr * foil_factor * night_factor * press_age_arr
                 + np.random.normal(0, WASTE_NOISE_STD, n)).clip(0.005, 0.18)

    waste_sheets   = (qty_sheets * waste_pct).astype(int)
    sheets_run     = qty_sheets + waste_sheets
    qty_cards      = qty_sheets * cards_per_sheet

    # Sheetfed 4/4 = 2 passes; perfecting = 1
    passes     = np.where(is_perfecting, 1, 2)
    run_time   = (sheets_run * passes / press_speed).round(2)
    total_time = ((makeready_time / 60) + run_time).round(2)

    # Ink
    ink_coverage = np.random.uniform(*INK_COVERAGE_RANGE, n)
    ink_lbs      = (sheets_run * ink_coverage * 4 * 2 / 50000).round(2)

    # Quality metrics
    quality_factor = np.where(is_night, NIGHT_QUALITY_FACTOR, 1.0) * press_age_arr
    delta_e_base   = np.array([DELTA_E_BASE[s] for s in stock_type])
    delta_e        = (np.random.exponential(delta_e_base, n) * quality_factor).clip(0.1, 9.0).round(2)

    reg_base       = np.array([REGISTER_BASE[s] for s in stock_type])
    register_err   = (np.random.exponential(reg_base, n) * quality_factor).clip(0.0, 5.0).round(3)

    dot_gain       = np.random.normal(21, 3, n).clip(10, 38).round(1)
    cut_dev        = (np.random.exponential(0.12, n) * foil_factor * quality_factor).clip(0.0, 1.5).round(3)
    foil_adhesion  = np.where(is_foil, np.random.normal(88, 8, n).clip(40, 100).round(1), np.nan)

    quality_pass = (
        (delta_e   < DELTA_E_REJECT) &
        (register_err < REGISTER_REJECT) &
        (dot_gain  < DOT_GAIN_REJECT) &
        (cut_dev   < CUT_DEVIATION_REJECT)
    ).astype(int)
    foil_fail    = np.where((stock_type == "Foil") & (foil_adhesion < FOIL_ADHESION_FAIL), 1, 0)
    quality_pass = np.where(foil_fail == 1, 0, quality_pass)
    rerun        = np.where(quality_pass == 0,
                            np.random.choice([0, 1], n, p=[1 - RERUN_PROBABILITY, RERUN_PROBABILITY]), 0)

    # Costs
    paper_cost_arr  = np.array([STOCK_COST_PER_MSF[s] for s in stock_type])
    paper_cost      = (sheets_run / 1000 * paper_cost_arr).round(2)

    rate_arr        = np.where(is_perfecting, PERFECTING_RATE_HR, SHEETFED_RATE_HR)
    press_cost      = (total_time * rate_arr).round(2)
    ink_cost        = (ink_lbs * INK_COST_PER_LB).round(2)
    plate_cost      = np.array([PLATE_COST[c] for c in ink_config], dtype=float)
    finishing_cost  = (sheets_run * np.array([FINISHING_COST_PER_SHEET[c] for c in ink_config])).round(2)
    total_cost      = (paper_cost + press_cost + ink_cost + plate_cost + finishing_cost).round(2)

    markup          = np.random.uniform(MARKUP_MIN, MARKUP_MAX, n)
    revenue         = (total_cost * markup).round(2)
    gross_profit    = (revenue - total_cost).round(2)
    gross_margin    = (gross_profit / revenue * 100).round(1)
    cost_per_card   = (total_cost / qty_cards).round(4)
    rev_per_card    = (revenue    / qty_cards).round(4)

    # Delivery
    due_dates      = [d + timedelta(days=int(random.gauss(LEAD_TIME_MEAN_DAYS, LEAD_TIME_STD_DAYS))) for d in dates]
    actual_del     = [d + timedelta(days=int(random.gauss(ACTUAL_TIME_MEAN, ACTUAL_TIME_STD)))       for d in dates]
    on_time        = [1 if a <= due else 0 for a, due in zip(actual_del, due_dates)]

    df = pd.DataFrame({
        "job_id":               [f"JOB-{str(i).zfill(5)}" for i in range(1, n + 1)],
        "job_date":             [d.strftime("%Y-%m-%d") for d in dates],
        "customer_id":          customers,
        "card_set_id":          card_sets,
        "shift":                shifts,
        "press":                presses,
        "press_type":           press_type_arr,
        "passes_required":      passes,
        "stock_type":           stock_type,
        "ink_config":           ink_config,
        "cards_per_sheet":      cards_per_sheet,
        "qty_sheets_ordered":   qty_sheets,
        "qty_cards_ordered":    qty_cards,
        "sheets_run":           sheets_run,
        "waste_sheets":         waste_sheets,
        "waste_pct":            (waste_pct * 100).round(2),
        "press_speed_sph":      press_speed,
        "makeready_time_min":   makeready_time,
        "run_time_hrs":         run_time,
        "total_press_time_hrs": total_time,
        "ink_lbs_used":         ink_lbs,
        "ink_coverage_pct":     (ink_coverage * 100).round(1),
        "color_delta_e":        delta_e,
        "register_error_thou":  register_err,
        "dot_gain_pct":         dot_gain,
        "foil_adhesion_score":  foil_adhesion,
        "cut_deviation_mm":     cut_dev,
        "quality_pass":         quality_pass,
        "rerun_required":       rerun,
        "paper_cost":           paper_cost,
        "press_cost":           press_cost,
        "ink_cost":             ink_cost,
        "plate_cost":           plate_cost,
        "finishing_cost":       finishing_cost,
        "total_cost":           total_cost,
        "revenue":              revenue,
        "gross_profit":         gross_profit,
        "gross_margin_pct":     gross_margin,
        "cost_per_card":        cost_per_card,
        "revenue_per_card":     rev_per_card,
        "on_time_delivery":     on_time,
    })

    return df


def print_summary(df):
    print("\n=== DATASET SUMMARY ===")
    print(f"Jobs:              {len(df):,}")
    print(f"Date range:        {df['job_date'].min()} → {df['job_date'].max()}")
    print(f"Total revenue:     ${df['revenue'].sum():>12,.0f}")
    print(f"Total gross profit:${df['gross_profit'].sum():>12,.0f}")
    print(f"Avg gross margin:  {df['gross_margin_pct'].mean():.1f}%")
    print(f"Quality pass rate: {df['quality_pass'].mean():.1%}")
    print(f"Rerun rate:        {df['rerun_required'].mean():.1%}")
    print(f"Rerun cost impact: ${df[df['rerun_required']==1]['total_cost'].sum():>12,.0f}")
    print(f"\n--- By Press ---")
    print(df.groupby(["press","press_type"])[
        ["waste_pct","quality_pass","makeready_time_min"]].mean().round(2))
    print(f"\n--- By Shift ---")
    print(df.groupby("shift")[["waste_pct","quality_pass","rerun_required"]].mean().round(3))
    print(f"\n--- By Stock ---")
    print(df.groupby("stock_type")[
        ["waste_pct","quality_pass","cost_per_card","gross_margin_pct"]].mean().round(3))


def export_excel(df, path):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("openpyxl not installed — skipping Excel export. Run: pip install openpyxl")
        return

    DARK_BLUE = "1F4E79"
    ALT_ROW   = "EBF3FA"
    WHITE_    = "FFFFFF"

    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = "Job Data"

    hfont = Font(bold=True, color=WHITE_, name="Arial", size=10)
    dfont = Font(name="Arial", size=10)
    hfill = PatternFill("solid", start_color=DARK_BLUE)
    afill = PatternFill("solid", start_color=ALT_ROW)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for ci, col in enumerate(df.columns, 1):
        c = ws.cell(row=1, column=ci, value=col)
        c.font = hfont; c.fill = hfill; c.alignment = center_align
        ws.column_dimensions[get_column_letter(ci)].width = 16

    for ri, row in enumerate(df.itertuples(index=False), 2):
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=ri, column=ci, value=val if str(val) != "nan" else None)
            c.font = dfont
            if ri % 2 == 0:
                c.fill = afill

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"
    wb.save(path)
    print(f"Excel saved:  {path}")


# ── MAIN ───────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    print("Generating dataset...")
    df = generate_dataset()
    print_summary(df)

    df.to_csv(OUTPUT_CSV, index=False)
    print(f"\nCSV saved:    {OUTPUT_CSV}")

    if EXPORT_EXCEL and OUTPUT_XLSX:
        export_excel(df, OUTPUT_XLSX)

    print("\nDone.")
